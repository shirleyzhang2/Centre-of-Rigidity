import os
import win32com.client
import openpyxl
import random
from openpyxl import *
import re
import time
import scipy
import numpy
from scipy.stats import norm
import datetime
import matplotlib.pyplot as plt

from util_param import *

def write_to_excel(wb,all_floors,save_loc):
    print('Writing all results to Excel...')
    filepath = save_loc + '/Results.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Floor #'
    ws['B1'] = 'load location (x)'
    ws['C1'] = 'Disp 1'
    ws['D1'] = 'Disp 2'

    iter = 2
    for i in range(len(all_floors)):
        floor_index = all_floors[i]
        for y_cord, disp in floor_index.items():
            ws['A'+str(iter)] = str(i+1)
            ws['B'+str(iter)] = str(y_cord)
            ws['C'+str(iter)] = str(disp[0])
            ws['D'+str(iter)] = str(disp[1])
            iter += 1
        iter += 1
    wb.save(filepath)

def get_displacement(SapModel,floor_elev):
    SapModel.SetPresentUnits(kip_in_F)
    [ret, number_nodes, all_node_names] = SapModel.PointObj.GetNameList()
    nodes = [None]*2
    
    #OPTIMIZE
    for node_name in all_node_names:
        [ret, x, y, z] = SapModel.PointObj.GetCoordCartesian(node_name,0,0,0)
        if x == 0 and y == 0 and z == floor_elev:
            nodes[0] = node_name
        if x == 0 and y == 12 and z == floor_elev:
            nodes[1] = node_name


    drift = [0]*3
    #Find displacements for nodes
    for i in range(len(nodes)):
        ret = SapModel.Results.JointDispl(nodes[i],0)
        max_and_min_disp = ret[7] # ret[7] is U1, U1 is a list of 2 values
        min_neg_disp = 0
        if len(max_and_min_disp) >= 2:
            min_neg_disp = max_and_min_disp[1]
        max_pos_disp = max_and_min_disp[0]
        if abs(max_pos_disp) >= abs(min_neg_disp):
            drift[i] = abs(max_pos_disp)
        elif abs(min_neg_disp) >= abs(max_pos_disp):
            drift[i] = abs(min_neg_disp)
        else:
            print('Could not find max drift for node {}'.format(i+1))
    return drift

print('\n--------------------------------------------------------')
print('Centre of Rigidity Finder by University of Toronto Seismic Design Team')
print('--------------------------------------------------------\n')

print('\nInitializing SAP2000 model...')
SapObject = win32com.client.Dispatch('SAP2000v15.SapObject') # create SAP2000 object
SapObject.ApplicationStart() # start SAP2000
SapModel = SapObject.SapModel # create SapModel Object
SapModel.InitializeNewModel() # initialize model

# CHANGE THE FILENAME TO THE PATH OF YOUR TOWER
FileName = "C:\Check Cr\Maher.sdb"
path='C:/Users/shirl/OneDrive - University of Toronto/Desktop/Seismic/Centre of Rigidity/Centre of Rigidity/temp.xlsx'
wb = load_workbook(path)
ret = SapModel.File.OpenFile(FileName) 
SapModel.SetPresentUnits(kip_in_F) #set imperial units for location of point force

#Variable initialization
floor_elev = [57.0] #[15.0,21.0,27.0,33.0,39.0,45.0,51.0,57.0] #change this list for total number of floors analyzed
cur_floor_num = 0
all_floors = []
mass_name = None

while cur_floor_num < len(floor_elev):
    all_nodes = {}
    cur_elev = floor_elev[cur_floor_num]
    y_cord = 0.0

    while y_cord <= 12.0:
        ret = SapModel.SetModelIsLocked(False) # Ensure unlocked
        print('Deleting and adding point load...')
        if mass_name is not None:
            ret = SapModel.PointObj.DeleteLoadForce(mass_name, 'DEAD') #delete point load
        [ret, mass_name] = SapModel.PointObj.AddCartesian(0,y_cord,cur_elev, MergeOff=False)  #Create the load node point
        if ret != 0:
            print('ERROR setting mass nodes on floor')

        #Assign masses to the mass nodes
        SapModel.SetPresentUnits(N_m_C) 
        mass_per_node = 0.2
        ret = SapModel.PointObj.SetLoadForce(mass_name, 'DEAD', [mass_per_node * 9.81, 0, 0, 0, 0, 0]) #Shaking in the x direcion!
        if ret[0] != 0:
            print('ERROR setting load on floor') # + str(floor_num)

        #Start analysis
        print('Computing...')
        SapModel.Analyze.RunAnalysis()
        print('Finished computing.')

        ret = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = SapModel.Results.Setup.SetCaseSelectedForOutput('DEAD', True)        
        
        all_nodes[y_cord] = get_displacement(SapModel,cur_elev)
       
        y_cord += 0.5 #Change this value for accuracy i.e. number of loads on each floor
    all_floors.append(all_nodes)
    cur_floor_num += 1 

SaveLoc = "C:/Users/shirl/OneDrive - University of Toronto/Desktop/Seismic/Centre of Rigidity/Centre of Rigidity"
write_to_excel(wb,all_floors,SaveLoc)

